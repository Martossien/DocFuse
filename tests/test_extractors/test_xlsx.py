"""Tests de l'extracteur XLSX.

CdC §8.3 — Chaque feuille, cellules non vides, ordre A1.
Feuille vide signalée. Nom de feuille en titre.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest

from docfuse.core.ocr.tesseract import TesseractEngine
from docfuse.extractors.xlsx import XlsxExtractor
from docfuse.models.file_status import FileStatus

_OCR_AVAILABLE = TesseractEngine().is_available()


def _xlsx_with_image(tmp_path: Path, name: str, image_bytes: bytes) -> Path:
    import openpyxl
    from openpyxl.drawing.image import Image as XlsxImage

    f = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feuil1"
    ws["A1"] = "texte avant image"
    img = XlsxImage(io.BytesIO(image_bytes))
    img.anchor = "C3"
    ws.add_image(img)
    wb.save(str(f))
    return f


def _red_square_png() -> bytes:
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGB", (50, 50), "red").save(buf, format="PNG")
    return buf.getvalue()


class TestXlsxExtractor:
    def test_extract_basic_sheet(self, tmp_path: Path) -> None:
        import openpyxl

        f = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Nom"
        ws["B1"] = "Valeur"
        ws["A2"] = "Test avec du texte"
        ws["B2"] = "42"
        wb.save(str(f))

        result = XlsxExtractor.extract(f, "test.xlsx")
        assert result.status is FileStatus.READY
        assert "Nom" in result.text
        assert "Valeur" in result.text
        assert "Data" in result.text  # Nom de feuille

    def test_empty_sheet_signaled(self, tmp_path: Path) -> None:
        import openpyxl

        f = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Vide"
        wb.save(str(f))

        result = XlsxExtractor.extract(f, "empty.xlsx")
        assert "[Feuille vide]" in result.text

    def test_accepts(self) -> None:
        assert XlsxExtractor.accepts(Path("test.xlsx")) is True
        assert XlsxExtractor.accepts(Path("test.csv")) is False

    def test_safe_extract_no_crash(self, tmp_path: Path) -> None:
        f = tmp_path / "nonexistent.xlsx"
        result = XlsxExtractor.safe_extract(f, "nonexistent.xlsx")
        assert result.status is FileStatus.ERROR

    def test_password_protected_gives_clear_error(self, tmp_path: Path) -> None:
        """D-089 : un .xlsx protégé par mot de passe à l'ouverture (conteneur
        OLE2, plus un ZIP) donnait un `BadZipFile` bas niveau, jamais un
        message disant à l'utilisateur que le fichier est protégé."""
        f = tmp_path / "protected.xlsx"
        f.write_bytes(bytes((0xD0, 0xCF, 0x11, 0xE0, 0xA1, 0xB1, 0x1A, 0xE1)) + b"\x00" * 500)

        result = XlsxExtractor.extract(f, "protected.xlsx")
        assert result.status is FileStatus.ERROR
        assert result.error_message is not None
        assert "mot de passe" in result.error_message.lower()

    def test_fixture_file(self) -> None:
        fixture = Path(__file__).resolve().parent.parent / "fixtures" / "sample.xlsx"
        if fixture.exists():
            result = XlsxExtractor.extract(fixture, "sample.xlsx")
            assert result.status is FileStatus.READY
            assert "Feuille1" in result.text

    def test_uncalculated_formula_text_is_recovered(self, tmp_path: Path) -> None:
        """D-076 : une formule jamais calculée (fichier généré par script,
        jamais ouvert dans Excel/LibreOffice — pas de valeur en cache dans
        le fichier) ne doit pas disparaître silencieusement. `data_only=True`
        renvoie None pour ces cellules, indistinguable d'une cellule
        réellement vide sans relire le classeur en data_only=False."""
        import openpyxl

        f = tmp_path / "formulas.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Total"
        ws["B1"] = "=1+1"  # jamais "calculée" : openpyxl n'évalue jamais les formules
        wb.save(str(f))

        result = XlsxExtractor.extract(f, "formulas.xlsx")
        assert result.status is FileStatus.READY
        assert "=1+1" in result.text

    def test_incorrect_declared_dimension_does_not_truncate_rows(self, tmp_path: Path) -> None:
        """D-084 : en mode read_only, openpyxl fait confiance à l'élément XML
        <dimension> déclaré par le fichier plutôt que de scanner le contenu
        réel. Un générateur tiers qui écrit une dimension trop petite (bug
        connu et documenté par openpyxl lui-même) faisait tronquer
        silencieusement les lignes en fin de feuille."""
        import zipfile

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=f"Row{i}")
        f = tmp_path / "lied_dimension.xlsx"
        wb.save(str(f))

        # Mentir sur la dimension déclarée : A1:A10 -> A1:A3.
        with zipfile.ZipFile(str(f)) as zin:
            items = zin.infolist()
            contents = {item.filename: zin.read(item.filename) for item in items}
        sheet_xml = contents["xl/worksheets/sheet1.xml"].decode("utf-8")
        sheet_xml = sheet_xml.replace('<dimension ref="A1:A10"/>', '<dimension ref="A1:A3"/>')
        contents["xl/worksheets/sheet1.xml"] = sheet_xml.encode("utf-8")
        with zipfile.ZipFile(str(f), "w", zipfile.ZIP_DEFLATED) as zout:
            for item in items:
                zout.writestr(item, contents[item.filename])

        result = XlsxExtractor.extract(f, "lied_dimension.xlsx")
        assert result.status is FileStatus.READY
        for i in range(1, 11):
            assert f"Row{i}" in result.text

    def test_merged_cells_value_is_propagated(self, tmp_path: Path) -> None:
        """D-085 : seule la cellule en haut à gauche d'une plage fusionnée
        porte une valeur (comportement Excel normal) — `ReadOnlyWorksheet`
        n'expose même pas `merged_cells`. Sans propagation, une ligne dont
        le titre fusionné s'étale sur plusieurs colonnes perd tout contexte
        pour les cellules "creuses" qui suivent."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Titre fusionne"
        ws.merge_cells("A1:C1")
        ws["A2"] = "Alpha"
        ws["B2"] = "Beta"
        ws["C2"] = "Gamma"
        f = tmp_path / "merged_horizontal.xlsx"
        wb.save(str(f))

        result = XlsxExtractor.extract(f, "merged_horizontal.xlsx")
        assert result.status is FileStatus.READY
        assert "Titre fusionne | Titre fusionne | Titre fusionne" in result.text

    def test_merged_cells_vertical_value_is_propagated(self, tmp_path: Path) -> None:
        """D-085, fusion verticale (sur plusieurs lignes plutôt que colonnes)."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Section"
        ws.merge_cells("A1:A3")
        ws["B1"] = "L1"
        ws["B2"] = "L2"
        ws["B3"] = "L3"
        f = tmp_path / "merged_vertical.xlsx"
        wb.save(str(f))

        result = XlsxExtractor.extract(f, "merged_vertical.xlsx")
        assert result.status is FileStatus.READY
        assert "Section | L1" in result.text
        assert "Section | L2" in result.text
        assert "Section | L3" in result.text

    def test_truly_empty_cell_stays_empty(self, tmp_path: Path) -> None:
        """Non-régression : une cellule réellement vide (pas une formule non
        calculée) ne doit pas récupérer de faux marqueur de formule."""
        import openpyxl

        f = tmp_path / "blank_cell.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Label"
        ws["A3"] = "AfterGap"
        wb.save(str(f))

        result = XlsxExtractor.extract(f, "blank_cell.xlsx")
        assert result.status is FileStatus.READY
        assert "formule" not in result.text

    def test_embedded_image_export_creates_embedded_images(self, tmp_path: Path) -> None:
        """D-093 : export actif -> l'image est capturée avec un nom explicite
        (feuille incluse) et un tag `[[IMAGE: ...]]` ajouté en fin de feuille."""
        f = _xlsx_with_image(tmp_path, "with_image.xlsx", _red_square_png())

        result = XlsxExtractor.extract(f, "with_image.xlsx", extract_images=True)
        assert result.status is FileStatus.READY
        assert len(result.embedded_images) == 1
        image = result.embedded_images[0]
        assert image.filename.startswith("with_image__sheet_Feuil1__img1")
        assert image.data
        assert f"[[IMAGE: {image.filename}]]" in result.text

    def test_embedded_image_export_disabled_by_default(self, tmp_path: Path) -> None:
        """D-093 : sans `extract_images`, aucune image capturée (non-régression)."""
        f = _xlsx_with_image(tmp_path, "with_image2.xlsx", _red_square_png())

        result = XlsxExtractor.extract(f, "with_image2.xlsx")
        assert result.embedded_images == []
        assert "[[IMAGE" not in result.text

    @pytest.mark.skipif(not _OCR_AVAILABLE, reason="Tesseract non installé")
    def test_embedded_image_ocr_extracts_text_automatically(self, tmp_path: Path) -> None:
        """D-093 : OCR automatique des images intégrées Excel, sans avoir
        besoin d'activer l'export."""
        from PIL import Image, ImageDraw

        img = Image.new("RGB", (300, 80), "white")
        draw = ImageDraw.Draw(img)
        draw.text((10, 30), "Bonjour Excel", fill="black")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        f = _xlsx_with_image(tmp_path, "with_text_image.xlsx", buf.getvalue())

        result = XlsxExtractor.extract(f, "with_text_image.xlsx")
        assert "onjour" in result.text or "Excel" in result.text
        assert result.embedded_images == []


def _xlsx_with_unsupported_extension(tmp_path: Path) -> Path:
    """Classeur portant une extension OOXML qu'openpyxl ne sait pas modéliser.

    C'est le cas réel du serveur de production : openpyxl émet alors
    « Data Validation extension is not supported and will be removed » à
    chaque lecture de feuille. La validation de données x14 ne s'écrit pas
    avec openpyxl — le bloc `<extLst>` est injecté directement dans le XML de
    la feuille, comme le fait Excel.
    """
    import zipfile

    import openpyxl

    plain = tmp_path / "_plain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Client"
    ws["B1"] = "Montant"
    ws["A2"] = "ACME"
    ws["B2"] = 42
    wb.save(str(plain))

    ext_block = (
        b'<extLst><ext uri="{CCE6A557-97BC-4b89-ADB6-D9C93CAAB3DF}" '
        b'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
        b'<x14:dataValidations count="0"/></ext></extLst>'
    )
    f = tmp_path / "validation.xlsx"
    with zipfile.ZipFile(plain) as zin, zipfile.ZipFile(f, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(b"</worksheet>", ext_block + b"</worksheet>")
            zout.writestr(item, data)
    return f


class TestOpenpyxlWarnings:
    """D-105 : le bruit openpyxl (« … extension is not supported and will be
    removed ») remontait dans la console de l'exécutable et inquiétait
    l'utilisateur, alors qu'il est sans effet sur le texte extrait."""

    def test_fixture_really_triggers_the_warning(self, tmp_path: Path) -> None:
        """Garde-fou du test suivant : sans le filtre, openpyxl avertit bien."""
        import warnings

        import openpyxl

        f = _xlsx_with_unsupported_extension(tmp_path)
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            openpyxl.load_workbook(str(f))
        assert any("extension is not supported" in str(w.message) for w in caught)

    def test_extraction_emits_no_openpyxl_warning(self, tmp_path: Path) -> None:
        import warnings

        from docfuse.extractors.xlsx import silence_openpyxl_warnings

        f = _xlsx_with_unsupported_extension(tmp_path)
        with warnings.catch_warnings(record=True) as caught:
            # Le filtre est posé par le point d'entrée applicatif
            # (`cli.main` / `gui.launch`, D-106) : on reproduit ici cet état.
            silence_openpyxl_warnings()
            result = XlsxExtractor.extract(f, "validation.xlsx")

        assert result.status == FileStatus.READY
        assert "ACME" in result.text
        assert [str(w.message) for w in caught if w.category is UserWarning] == []


class TestOpenpyxlWarningsAreNotAGlobalSideEffect:
    """D-106 : le filtre était posé **à l'import du module**, donc sur le
    processus hôte, sans opt-out. Une application qui avait choisi
    `-W error::UserWarning` le perdait en silence du seul fait d'importer
    DocFuse ; la place d'une politique d'avertissements est le point d'entrée
    applicatif."""

    def test_importing_the_module_installs_no_filter(self) -> None:
        import subprocess
        import sys

        # Processus neuf : `warnings.filters` intact au moment de l'import.
        code = (
            "import warnings, sys;"
            "before = list(warnings.filters);"
            "import docfuse.extractors.xlsx;"
            "print(warnings.filters == before)"
        )
        proc = subprocess.run(
            [sys.executable, "-c", code], capture_output=True, text=True, check=True
        )
        assert proc.stdout.strip() == "True", (
            "importer docfuse.extractors.xlsx modifie warnings.filters du "
            f"processus hôte : {proc.stdout!r}"
        )

    def test_host_policy_w_error_survives_the_import(self) -> None:
        """La preuve côté utilisateur : une application hôte qui a posé
        `-W error::UserWarning` doit garder son choix après un simple `import
        docfuse`. Avant D-106, le filtre posé à l'import passait devant (il est
        inséré **en tête** de `warnings.filters`) et l'avertissement openpyxl
        était avalé, sans opt-out."""
        import subprocess
        import sys

        # `warn_explicit(..., module=...)` reproduit un avertissement émis
        # depuis openpyxl sans avoir à ouvrir un vrai classeur.
        code = (
            "import docfuse.extractors.xlsx, warnings;"
            "warnings.warn_explicit("
            "'Unknown extension is not supported and will be removed',"
            " UserWarning, 'openpyxl/reader/excel.py', 1, module='openpyxl.reader.excel');"
            "print('AVALE')"
        )
        proc = subprocess.run(
            [sys.executable, "-W", "error::UserWarning", "-c", code],
            capture_output=True,
            text=True,
        )
        assert proc.returncode != 0, f"le -W error de l'hôte a été neutralisé : {proc.stdout!r}"
        assert "UserWarning" in proc.stderr

    def test_entry_points_install_the_filter(self) -> None:
        """`cli.main` et `gui.launch` appellent bien la fonction publique."""
        from pathlib import Path

        root = Path(__file__).resolve().parent.parent.parent / "src" / "docfuse"
        for name in ("cli.py", "gui/app.py"):
            source = (root / name).read_text("utf-8")
            assert "silence_openpyxl_warnings()" in source, (
                f"{name} ne pose plus la politique d'avertissements openpyxl"
            )

    def test_filter_regex_is_anchored_on_openpyxl(self) -> None:
        """La regex `module=` est compilée puis `.match()`ée : non ancrée,
        elle couvrait aussi `openpyxl_autre.chose`."""
        import re

        from docfuse.extractors.xlsx import _OPENPYXL_MODULE_RE

        pattern = re.compile(_OPENPYXL_MODULE_RE)
        assert pattern.match("openpyxl")
        assert pattern.match("openpyxl.worksheet._reader")
        assert not pattern.match("openpyxl_autre.chose")
