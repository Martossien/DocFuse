"""Extracteur XLSX : .xlsx.

CdC §8.3 — Chaque feuille, cellules non vides, ordre A1…
Feuille vide signalée. Nom de feuille en titre.
"""

from __future__ import annotations

import logging
import posixpath
import re
import warnings
import xml.etree.ElementTree as ET
import zipfile
from contextlib import closing
from pathlib import Path

from docfuse.core.embedded_images import ImageBatch, build_image_tag
from docfuse.core.ocr.registry import resolve_ocr_engine
from docfuse.core.registry import register
from docfuse.extractors.base import Extractor, container_guard, error_result, file_type_for
from docfuse.models.extraction_result import ExtractedFile
from docfuse.models.file_status import FileStatus

logger = logging.getLogger(__name__)

_MERGE_CELL_REF_RE = re.compile(r'<mergeCell ref="([^"]+)"')


# Le filtre vise le module émetteur — `openpyxl` lui-même ou un sous-module.
# D-106 : la regex était `"openpyxl"`, non ancrée : `warnings` la compile puis
# appelle `.match()`, donc `openpyxl_autre.chose` était couvert aussi.
_OPENPYXL_MODULE_RE = r"openpyxl(\.|$)"


def silence_openpyxl_warnings() -> None:
    """Neutralise les `UserWarning` d'openpyxl (D-105) — **API publique**.

    openpyxl avertit sur chaque fonctionnalité de classeur qu'il ne sait pas
    représenter en mémoire (« Data Validation extension is not supported and
    will be removed », « Conditional Formatting extension… », « Unknown
    extension… », en-têtes non conformes…). DocFuse ne réécrit **jamais** un
    classeur : il n'en lit que le texte, donc ces avertissements n'ont
    aucune conséquence sur la sortie — mais ils remontaient en masse dans la
    console de l'exécutable et inquiétaient l'utilisateur.

    Volontairement **pas** un `warnings.catch_warnings()` autour de chaque
    appel : ce dernier mute l'état global de `warnings` et n'est pas
    thread-safe — il fuirait entre les fichiers traités en parallèle par le
    `ThreadPoolExecutor` de l'orchestrateur (masquant des avertissements
    d'autres extracteurs, ou laissant passer ceux-ci selon l'entrelacement).

    D-106 : ce filtre n'est **plus posé à l'import du module**. Poser un
    filtre global depuis l'import d'une bibliothèque est un effet de bord sur
    le processus hôte, sans opt-out : il s'ajoute en tête de
    `warnings.filters` et une application qui avait choisi
    `-W error::UserWarning` perdait ce choix en silence, du seul fait
    d'importer DocFuse. La place d'une politique d'avertissements est le
    **point d'entrée applicatif** : `cli.main()` et `gui.launch()` l'appellent
    (voir aussi la note d'intégration côté docia). Une bibliothèque appelante
    décide elle-même, ou ne l'appelle pas.

    Le filtre vise le module émetteur (`openpyxl` et ses sous-modules) et la
    seule catégorie `UserWarning` : rien d'autre n'est masqué. Idempotent —
    `warnings.filterwarnings` dédoublonne les entrées identiques.
    """
    warnings.filterwarnings("ignore", category=UserWarning, module=_OPENPYXL_MODULE_RE)


@register(".xlsx")
class XlsxExtractor(Extractor):
    """Extracteur XLSX via openpyxl."""

    @classmethod
    def accepts(cls, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    @classmethod
    def extract(cls, path: Path, relative_path: str, extract_images: bool = False) -> ExtractedFile:
        try:
            # D-089/D-093 : fichier protégé par mot de passe (conteneur OLE2)
            # ou « bombe zip » — garde partagée entre conteneurs (D-099).
            guard = container_guard(path, relative_path)
            if guard is not None:
                return guard

            # D-093 : OCR/export des images intégrées, seulement si utile
            # (export demandé ou OCR disponible) — sinon zéro coût ajouté.
            # D-098 : jetons collectés par feuille, OCR parallèle unique.
            batch = ImageBatch(resolve_ocr_engine(), extract_images)
            sheets: list[tuple[str, str, list[str]]] = []

            from openpyxl import load_workbook

            # D-096 : `closing()` — les classeurs read_only gardent l'archive
            # ZIP ouverte jusqu'à `.close()`, qui n'était atteint que sur le
            # chemin nominal : toute exception dans la boucle des feuilles
            # laissait deux archives ouvertes jusqu'au GC.
            wb = closing(load_workbook(str(path), read_only=True, data_only=True))
            # D-076 : data_only=True renvoie None pour une formule jamais
            # calculée (fichier généré par script, jamais ouvert dans
            # Excel/LibreOffice — pas de valeur en cache dans le fichier).
            # Sans ce second classeur (data_only=False, qui donne le TEXTE
            # de la formule), la cellule paraît vide sans aucune trace
            # qu'un calcul existait — perte silencieuse de colonnes de
            # totaux entières sur des exports automatisés (ERP/BI).
            wb_formulas = closing(load_workbook(str(path), read_only=True, data_only=False))
            parts: list[str] = []

            with wb as wb_values, wb_formulas as wb_f, zipfile.ZipFile(str(path)) as zf:
                names = set(zf.namelist())
                sheet_count = len(wb_values.sheetnames)
                for sheet in wb_values.sheetnames:
                    ws = wb_values[sheet]
                    # D-096 : une feuille graphique (`Chartsheet`, onglet
                    # "Graphique1") est listée dans `sheetnames` mais n'a ni
                    # cellules ni `reset_dimensions` → `AttributeError` qui
                    # mettait TOUT le classeur en ERROR, données perdues.
                    # Signalée explicitement, jamais silencieuse.
                    if not hasattr(ws, "iter_rows"):
                        sheets.append((sheet, "[Feuille graphique — pas de cellules]", []))
                        continue
                    # D-098 : le XML brut de la feuille est lu une fois et
                    # sert trois usages (plages fusionnées D-085, présence de
                    # formules, images). Sans aucun élément `<f>`, le second
                    # classeur (formules, D-076) n'est pas consulté pour cette
                    # feuille : il ne peut renvoyer que `None` pour les cellules
                    # vides — sortie identique, deux parses complets évités.
                    merge_path = getattr(ws, "_worksheet_path", "").lstrip("/")
                    sheet_xml = zf.read(merge_path) if merge_path in names else b""
                    has_formulas = b"<f>" in sheet_xml or b"<f " in sheet_xml or b"<f/" in sheet_xml
                    ws_formulas = wb_f[sheet] if has_formulas and sheet in wb_f.sheetnames else None
                    if ws_formulas is not None and not hasattr(ws_formulas, "iter_rows"):
                        ws_formulas = None

                    # D-084 : en mode read_only, openpyxl fait confiance à
                    # l'élément XML <dimension> déclaré par le fichier
                    # plutôt que de scanner le contenu réel. Certains
                    # générateurs tiers écrivent une dimension incorrecte
                    # (trop petite) — iter_rows() tronque alors
                    # silencieusement les lignes/colonnes en fin de
                    # feuille, sans erreur. On force un vrai recalcul avant
                    # de lire. `calculate_dimension(force=True)` lève
                    # `UnboundLocalError` sur une feuille réellement vide
                    # (bug openpyxl : `cell` jamais assignée dans sa
                    # boucle) — sans conséquence, `has_data` reste `False`.
                    for candidate in (ws, ws_formulas):
                        if candidate is None:
                            continue
                        try:
                            candidate.reset_dimensions()
                            candidate.calculate_dimension(force=True)
                        except UnboundLocalError:
                            pass

                    rows_text: list[str] = []
                    has_data = False

                    formula_rows = (
                        ws_formulas.iter_rows(values_only=True)
                        if ws_formulas is not None
                        else iter(())
                    )
                    grid: list[list[str]] = []
                    for row in ws.iter_rows(values_only=True):
                        formula_row = next(formula_rows, ())
                        cells: list[str] = []
                        for idx, c in enumerate(row):
                            if c is not None:
                                cells.append(str(c))
                                continue
                            formula = formula_row[idx] if idx < len(formula_row) else None
                            if isinstance(formula, str) and formula.startswith("="):
                                cells.append(f"[formule non calculée: {formula}]")
                            else:
                                # BUG FIX: data_only=True retourne aussi None pour une
                                # cellule réellement vide. On la marque comme vide.
                                cells.append("")
                        grid.append(cells)

                    # D-085 : seule la cellule en haut à gauche d'une plage
                    # fusionnée porte une valeur — les autres sont vides
                    # (comportement Excel normal, pas un bug openpyxl).
                    # Sans propager cette valeur, une ligne dont le titre
                    # fusionné s'étale sur plusieurs colonnes/lignes perd
                    # tout contexte pour les cellules "creuses" qui suivent
                    # — très fréquent dans les tableaux "présentables"
                    # (rapports, tableaux de bord).
                    if sheet_xml:
                        _apply_merged_cells(grid, _merge_ranges(sheet_xml))

                    for cells in grid:
                        if any(c.strip() for c in cells):
                            has_data = True
                            rows_text.append(" | ".join(cells))

                    sheet_text = "\n".join(rows_text) if has_data else "[Feuille vide]"

                    # D-093 : images intégrées ancrées sur cette feuille —
                    # regroupées en fin de feuille plutôt qu'à la cellule
                    # d'ancrage exacte (voir docstring de `_extract_sheet_images`).
                    tokens = (
                        _extract_sheet_images(zf, names, merge_path, sheet, relative_path, batch)
                        if sheet_xml and batch.active
                        else []
                    )
                    sheets.append((sheet, sheet_text, tokens))

            batch.run()
            for sheet, sheet_text, tokens in sheets:
                markers = batch.apply(tokens)
                if markers:
                    sheet_text += "\n\n" + "\n\n".join(markers)
                parts.append(f"### Feuille : {sheet}\n\n{sheet_text}")
            embedded_images = batch.images

            text = "\n\n".join(parts)

            return ExtractedFile(
                path=path,
                relative_path=relative_path,
                extension=file_type_for(path),
                file_type=file_type_for(path),
                size_bytes=path.stat().st_size,
                text=text,
                status=FileStatus.READY,
                page_count=sheet_count,
                embedded_images=embedded_images,
            )
        except Exception as exc:
            logger.exception("Erreur extraction XLSX %s", path)
            return error_result(path, relative_path, exc)


def _merge_ranges(sheet_xml: bytes) -> list[tuple[int, int, int, int]]:
    """Plages fusionnées `(min_col, min_row, max_col, max_row)` d'une feuille.

    Lu directement dans le XML de la feuille : `ReadOnlyWorksheet` (mode
    `read_only=True`, utilisé partout dans cet extracteur) n'expose pas
    `merged_cells` du tout (D-085) — seul le classeur normal l'expose,
    mais le charger entièrement en mémoire annulerait l'intérêt du mode
    `read_only` pour de gros fichiers.
    """
    from openpyxl.utils import range_boundaries

    xml = sheet_xml.decode("utf-8", errors="replace")

    ranges: list[tuple[int, int, int, int]] = []
    for ref in _MERGE_CELL_REF_RE.findall(xml):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        except ValueError:
            continue
        if None in (min_col, min_row, max_col, max_row):
            continue
        ranges.append((min_col, min_row, max_col, max_row))
    return ranges


def _extract_sheet_images(
    zf: zipfile.ZipFile,
    names: set[str],
    sheet_path: str,
    sheet_name: str,
    relative_path: str,
    batch: ImageBatch,
) -> list[str]:
    """Jetons des images intégrées ancrées sur une feuille XLSX (D-093, D-098).

    `openpyxl` en mode `read_only=True` (obligatoire ici pour supporter de
    gros classeurs) n'expose pas du tout les dessins/images — on lit donc
    le XML brut du ZIP directement, en suivant la même chaîne de relations
    OOXML que python-pptx/python-docx font automatiquement pour DOCX/PPTX :
    `sheetN.xml` → `_rels/sheetN.xml.rels` (relation "drawing") →
    `drawingM.xml` (ancres `oneCellAnchor`/`twoCellAnchor`, chacune avec un
    `<a:blip r:embed="rIdY">`) → `_rels/drawingM.xml.rels` (relation
    "image") → fichier média réel.

    Les marqueurs sont regroupés en fin de feuille plutôt qu'ancrés à la
    cellule exacte : la position (ligne, colonne) de l'ancre XML ne
    correspond pas forcément à une ligne "avec données" au sens du texte
    déjà généré (tableau pipe par ligne non vide) — une fausse précision
    d'ancrage serait trompeuse.

    Returns:
        Jetons de position (un par image, dans l'ordre des ancres), à
        résoudre via `batch.apply` après `batch.run`. N'échoue jamais :
        toute relation manquante ou XML illisible donne une liste vide.
    """
    tokens: list[str] = []

    sheet_rels_path = posixpath.join(
        posixpath.dirname(sheet_path), "_rels", posixpath.basename(sheet_path) + ".rels"
    )
    if sheet_rels_path not in names:
        return tokens

    try:
        drawing_target = _find_target_by_type(zf.read(sheet_rels_path), "/drawing")
    except Exception:
        return tokens
    if not drawing_target:
        return tokens

    drawing_path = _resolve_rel_target(sheet_path, drawing_target)
    if drawing_path not in names:
        return tokens

    drawing_rels_path = posixpath.join(
        posixpath.dirname(drawing_path), "_rels", posixpath.basename(drawing_path) + ".rels"
    )
    try:
        drawing_rels = (
            _parse_relationships(zf.read(drawing_rels_path)) if drawing_rels_path in names else {}
        )
        image_rids = _parse_drawing_images(zf.read(drawing_path))
    except Exception:
        logger.warning("Lecture du dessin %s échouée", drawing_path, exc_info=True)
        return tokens

    count = 0
    for rid in image_rids:
        target = drawing_rels.get(rid)
        if not target:
            continue
        media_path = _resolve_rel_target(drawing_path, target)
        try:
            data = zf.read(media_path)
        except KeyError:
            continue

        count += 1
        ext = media_path.rsplit(".", 1)[-1] if "." in media_path else "png"
        tokens.append(
            batch.add(build_image_tag(relative_path, f"sheet_{sheet_name}", count, ext), data)
        )

    return tokens


def _resolve_rel_target(owner_part: str, target: str) -> str:
    """Résout un `Target` de relation OOXML (absolu `/xl/...` ou relatif
    `../drawings/...`) en chemin ZIP sans slash initial."""
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = posixpath.dirname(owner_part)
    return posixpath.normpath(posixpath.join(base_dir, target))


def _find_target_by_type(rels_xml: bytes, type_suffix: str) -> str | None:
    """Premier `Target` d'un fichier `.rels` dont le `Type` se termine par
    `type_suffix` (ex: "/drawing")."""
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return None
    for rel in root:
        if (rel.get("Type") or "").endswith(type_suffix):
            return rel.get("Target")
    return None


def _parse_relationships(rels_xml: bytes) -> dict[str, str]:
    """Table `Id -> Target` complète d'un fichier `.rels`."""
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return {}
    return {rel.get("Id", ""): rel.get("Target", "") for rel in root}


def _parse_drawing_images(drawing_xml: bytes) -> list[str]:
    """`rId` (attribut `r:embed`) de chaque image d'un `drawingN.xml`, dans
    l'ordre d'apparition des ancres (`oneCellAnchor`/`twoCellAnchor`)."""
    try:
        root = ET.fromstring(drawing_xml)
    except ET.ParseError:
        return []
    rids: list[str] = []
    for anchor in root:
        if not (anchor.tag.endswith("}oneCellAnchor") or anchor.tag.endswith("}twoCellAnchor")):
            continue
        for el in anchor.iter():
            if el.tag.endswith("}blip"):
                rid = next((v for k, v in el.attrib.items() if k.endswith("}embed")), None)
                if rid:
                    rids.append(rid)
    return rids


def _apply_merged_cells(grid: list[list[str]], ranges: list[tuple[int, int, int, int]]) -> None:
    """Propage la valeur de la cellule en haut à gauche de chaque plage
    fusionnée à toutes les autres cellules de cette plage, en mutant `grid`
    en place (`grid[ligne][colonne]`, 0-indexé ; `ranges` est 1-indexé,
    convention Excel/openpyxl)."""
    for min_col, min_row, max_col, max_row in ranges:
        r0, c0 = min_row - 1, min_col - 1
        if r0 >= len(grid) or c0 >= len(grid[r0]):
            continue
        top_left_value = grid[r0][c0]
        if not top_left_value:
            continue
        for r in range(min_row - 1, max_row):
            if r >= len(grid):
                break
            row_cells = grid[r]
            for c in range(min_col - 1, max_col):
                if c >= len(row_cells) or (r, c) == (r0, c0):
                    continue
                if not row_cells[c]:
                    row_cells[c] = top_left_value
